from pathlib import Path

from openpyxl import Workbook

from openpyxl.styles import Font

from openpyxl.styles import PatternFill

from openpyxl.styles import Alignment


# =====================================================
# GENERAR LISTA DE EDICIÓN
# =====================================================

def generate_edit_list(

    clips,

    output_folder,

    filename,

    opponent

):

    output_folder = Path(

        output_folder

    )

    output_folder.mkdir(

        parents=True,

        exist_ok=True

    )

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Lista de edición"

    # =====================================================
    # CABECERA
    # =====================================================

    headers = [

        "Orden",

        "Título",

        "Categoría",

        "Informe",

        "Rival",

        "URL del vídeo",

        "Duración",

        "Observaciones"

    ]

    header_fill = PatternFill(

        fill_type="solid",

        start_color="1F4E78",

        end_color="1F4E78"

    )

    for col, header in enumerate(

        headers,

        start=1

    ):

        cell = sheet.cell(

            row=1,

            column=col

        )

        cell.value = header

        cell.font = Font(

            bold=True,

            color="FFFFFF"

        )

        cell.fill = header_fill

        cell.alignment = Alignment(

            horizontal="center",

            vertical="center"

        )

    # =====================================================
    # CLIPS
    # =====================================================

    for row, clip in enumerate(

        clips,

        start=2

    ):

        sheet.cell(

            row=row,

            column=1

        ).value = row - 1

        sheet.cell(

            row=row,

            column=2

        ).value = clip.get(

            "title",

            ""

        )

        sheet.cell(

            row=row,

            column=3

        ).value = clip.get(

            "category",

            ""

        )

        sheet.cell(

            row=row,

            column=4

        ).value = clip.get(

            "report_id",

            ""

        )

        sheet.cell(

            row=row,

            column=5

        ).value = opponent

        sheet.cell(

            row=row,

            column=6

        ).value = clip.get(

            "video_url",

            ""

        )

        sheet.cell(

            row=row,

            column=7

        ).value = clip.get(

            "duration",

            ""

        )

        sheet.cell(

            row=row,

            column=8

        ).value = ""

    # =====================================================
    # AJUSTAR COLUMNAS
    # =====================================================

    widths = {

        "A": 10,

        "B": 40,

        "C": 25,

        "D": 20,

        "E": 25,

        "F": 80,

        "G": 15,

        "H": 40

    }

    for column, width in widths.items():

        sheet.column_dimensions[

            column

        ].width = width

    # =====================================================
    # GUARDAR
    # =====================================================

    output_file = (

        output_folder

        /

        f"{filename}.xlsx"

    )

    workbook.save(

        output_file

    )

    return str(

        output_file

    )